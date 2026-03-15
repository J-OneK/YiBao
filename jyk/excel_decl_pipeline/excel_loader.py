"""Excel workbook loading and used-range trimming."""

from __future__ import annotations

import os
from typing import Dict, List, Optional, Tuple

from .models import CellModel, SheetModel, WorkbookModel

DEFAULT_DPI = 96
DEFAULT_ROW_PT = 13.0
DEFAULT_COL_WIDTH = 8.0  # in Excel character-width unit
CHAR_WIDTH_PX = 7
MIN_COL_PX = 30
MIN_ROW_PX = 16


def points_to_pixels(points: float, dpi: int = DEFAULT_DPI) -> int:
    return max(MIN_ROW_PX, int(round(points * dpi / 72.0)))


def col_width_to_pixels(width_chars: float) -> int:
    # Approximation that stays close to Excel's rendered widths.
    return max(MIN_COL_PX, int(round(width_chars * CHAR_WIDTH_PX + 5)))


def _is_non_empty(value: str) -> bool:
    return bool(value and str(value).strip())


def load_workbook_model(excel_path: str) -> WorkbookModel:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    ext = os.path.splitext(excel_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        sheets = _load_xlsx(excel_path)
    elif ext == ".xls":
        sheets = _load_xls(excel_path)
    else:
        raise ValueError(f"Unsupported Excel format: {ext}")

    if not sheets:
        raise ValueError("No visible non-empty sheets found.")
    return WorkbookModel(source_path=excel_path, sheets=sheets)


def _load_xlsx(path: str) -> List[SheetModel]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: List[SheetModel] = []

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue

        min_row = ws.min_row or 1
        max_row = ws.max_row or 0
        min_col = ws.min_column or 1
        max_col = ws.max_column or 0
        if max_row < min_row or max_col < min_col:
            continue

        merge_map: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        for merged in ws.merged_cells.ranges:
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merge_map[(r, c)] = (
                        merged.min_row,
                        merged.min_col,
                        merged.max_row,
                        merged.max_col,
                    )

        n_rows = max_row - min_row + 1
        n_cols = max_col - min_col + 1
        row_heights = []
        col_widths = []

        for r in range(min_row, max_row + 1):
            row_dim = ws.row_dimensions.get(r)
            row_pt = row_dim.height if row_dim and row_dim.height else DEFAULT_ROW_PT
            row_heights.append(points_to_pixels(float(row_pt)))

        for c in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            col_dim = ws.column_dimensions.get(col_letter)
            width_chars = col_dim.width if col_dim and col_dim.width else DEFAULT_COL_WIDTH
            col_widths.append(col_width_to_pixels(float(width_chars)))

        cells: Dict[Tuple[int, int], CellModel] = {}
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        ):
            for c_idx, cell in enumerate(row):
                abs_r = min_row + r_idx
                abs_c = min_col + c_idx
                merged = merge_map.get((abs_r, abs_c))

                is_child = False
                span_r, span_c = 1, 1
                if merged:
                    mr0, mc0, mr1, mc1 = merged
                    is_child = abs_r != mr0 or abs_c != mc0
                    span_r = mr1 - mr0 + 1
                    span_c = mc1 - mc0 + 1

                val = "" if cell.value is None else str(cell.value)
                cells[(r_idx, c_idx)] = CellModel(
                    row=r_idx,
                    col=c_idx,
                    value=val,
                    merge_span_row=span_r,
                    merge_span_col=span_c,
                    is_merged_child=is_child,
                )

        sheet = SheetModel(
            name=ws.title,
            cells=cells,
            n_rows=n_rows,
            n_cols=n_cols,
            row_heights_px=row_heights,
            col_widths_px=col_widths,
        )
        trimmed = _trim_to_used_area(sheet)
        if trimmed is not None:
            sheets.append(trimmed)

    wb.close()
    return sheets


def _xls_cell_to_text(cell, wb) -> str:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_BLANK:
        return ""
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return str(cell.value).strip()
    if cell.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_BOOLEAN):
        num = float(cell.value)
        if num.is_integer():
            return str(int(num))
        return str(num)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(cell.value)
    return str(cell.value).strip() if cell.value is not None else ""


def _load_xls(path: str) -> List[SheetModel]:
    import xlrd

    wb = xlrd.open_workbook(path, formatting_info=True)
    sheets: List[SheetModel] = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.visibility != 0:
            continue
        if ws.nrows == 0 or ws.ncols == 0:
            continue

        merge_map: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        for r0, r1, c0, c1 in ws.merged_cells:
            for r in range(r0, r1):
                for c in range(c0, c1):
                    merge_map[(r, c)] = (r0, c0, r1 - 1, c1 - 1)

        row_heights: List[int] = []
        for r in range(ws.nrows):
            if r in ws.rowinfo_map and ws.rowinfo_map[r].height:
                row_pt = ws.rowinfo_map[r].height / 20.0
            else:
                row_pt = DEFAULT_ROW_PT
            row_heights.append(points_to_pixels(row_pt))

        col_widths: List[int] = []
        for c in range(ws.ncols):
            if c in ws.colinfo_map and ws.colinfo_map[c].width:
                width_chars = ws.colinfo_map[c].width / 256.0
            else:
                width_chars = DEFAULT_COL_WIDTH
            col_widths.append(col_width_to_pixels(width_chars))

        cells: Dict[Tuple[int, int], CellModel] = {}
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                merged = merge_map.get((r, c))
                is_child = False
                span_r, span_c = 1, 1
                if merged:
                    mr0, mc0, mr1, mc1 = merged
                    is_child = r != mr0 or c != mc0
                    span_r = mr1 - mr0 + 1
                    span_c = mc1 - mc0 + 1
                val = _xls_cell_to_text(ws.cell(r, c), wb)
                cells[(r, c)] = CellModel(
                    row=r,
                    col=c,
                    value=val,
                    merge_span_row=span_r,
                    merge_span_col=span_c,
                    is_merged_child=is_child,
                )

        sheet = SheetModel(
            name=ws.name,
            cells=cells,
            n_rows=ws.nrows,
            n_cols=ws.ncols,
            row_heights_px=row_heights,
            col_widths_px=col_widths,
        )
        trimmed = _trim_to_used_area(sheet)
        if trimmed is not None:
            sheets.append(trimmed)
    return sheets


def _trim_to_used_area(sheet: SheetModel) -> Optional[SheetModel]:
    min_row = sheet.n_rows
    min_col = sheet.n_cols
    max_row = -1
    max_col = -1

    for (r, c), cell in sheet.cells.items():
        if cell.is_merged_child:
            continue
        if _is_non_empty(cell.value):
            min_row = min(min_row, r)
            min_col = min(min_col, c)
            max_row = max(max_row, r + cell.merge_span_row - 1)
            max_col = max(max_col, c + cell.merge_span_col - 1)

    if max_row < 0 or max_col < 0:
        return None

    trimmed_cells: Dict[Tuple[int, int], CellModel] = {}
    for (r, c), cell in sheet.cells.items():
        if r < min_row or r > max_row or c < min_col or c > max_col:
            continue

        new_row = r - min_row
        new_col = c - min_col
        span_r = min(cell.merge_span_row, max_row - r + 1)
        span_c = min(cell.merge_span_col, max_col - c + 1)
        trimmed_cells[(new_row, new_col)] = CellModel(
            row=new_row,
            col=new_col,
            value=cell.value,
            merge_span_row=span_r,
            merge_span_col=span_c,
            is_merged_child=cell.is_merged_child,
        )

    return SheetModel(
        name=sheet.name,
        cells=trimmed_cells,
        n_rows=max_row - min_row + 1,
        n_cols=max_col - min_col + 1,
        row_heights_px=sheet.row_heights_px[min_row : max_row + 1],
        col_widths_px=sheet.col_widths_px[min_col : max_col + 1],
    )


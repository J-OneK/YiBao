"""
excel_to_text.py
把 Excel 文件的所有单元格文本提取出来，输出为易于大模型检索的纯文本。

格式示例（每行一个非空单元格，附坐标）：
    [Sheet1][A1] 货物名称
    [Sheet1][B1] 加拿大蒙特利尔
    ...

同时也会输出一个"所有词汇集合"文本，方便模型做模糊匹配。

依赖: openpyxl  xlrd==1.2.0
"""

import os
import re
import xlrd
import openpyxl
from typing import List, Tuple, Dict


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _col_letter(col_idx: int) -> str:
    """0-based 列号 → Excel 列字母, 例如 0->'A', 25->'Z', 26->'AA'"""
    result = ""
    n = col_idx + 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _cell_addr(row: int, col: int) -> str:
    """0-based (row, col) → 'A1' 格式"""
    return f"{_col_letter(col)}{row + 1}"


# ── xlsx 提取 ─────────────────────────────────────────────────────────────────

def _extract_xlsx(path: str) -> List[Tuple[str, str, str]]:
    """
    返回列表: [(sheet_name, cell_addr, cell_value), ...]
    只保留非空单元格，合并区域只取左上角那格。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if ws.max_row is None:
            continue

        min_r = ws.min_row or 1
        min_c = ws.min_column or 1

        # 合并区域：记录非左上角的格子，跳过它们
        merged_children = set()
        for m in ws.merged_cells.ranges:
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    if r != m.min_row or c != m.min_col:
                        merged_children.add((r, c))

        for row in ws.iter_rows(min_row=min_r, max_row=ws.max_row,
                                  min_col=min_c, max_col=ws.max_column):
            for cell in row:
                if (cell.row, cell.column) in merged_children:
                    continue
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                if not val:
                    continue
                ri = cell.row - min_r      # 0-based
                ci = cell.column - min_c   # 0-based
                records.append((ws.title, _cell_addr(ri, ci), val))

    wb.close()
    return records


# ── xls 提取 ──────────────────────────────────────────────────────────────────

def _extract_xls(path: str) -> List[Tuple[str, str, str]]:
    wb = xlrd.open_workbook(path, formatting_info=True)
    records = []

    for si in range(wb.nsheets):
        ws = wb.sheet_by_index(si)
        if ws.visibility != 0:
            continue

        # 合并区域子格
        merged_children = set()
        for (r0, r1, c0, c1) in ws.merged_cells:
            for r in range(r0, r1):
                for c in range(c0, c1):
                    if r != r0 or c != c0:
                        merged_children.add((r, c))

        for ri in range(ws.nrows):
            for ci in range(ws.ncols):
                if (ri, ci) in merged_children:
                    continue
                cell = ws.cell(ri, ci)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    val = cell.value.strip()
                elif cell.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_BOOLEAN):
                    num = cell.value
                    val = str(int(num)) if num == int(num) else str(num)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        import datetime
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        val = dt.strftime("%Y-%m-%d")
                    except Exception:
                        val = str(cell.value)
                else:
                    val = str(cell.value).strip() if cell.value else ""

                if not val:
                    continue
                records.append((ws.name, _cell_addr(ri, ci), val))

    return records


# ── 主提取函数 ────────────────────────────────────────────────────────────────

def extract_excel_text(
    excel_path: str,
    output_path: str | None = None,
) -> str:
    """
    从 Excel 文件提取所有单元格文本，返回格式化字符串。
    如果指定 output_path，同时写入文件。

    返回文本格式:
        ==== Sheet: Sheet1 ====
        [A1] 货物名称
        [B1] 加拿大蒙特利尔
        ...
        ---- 词汇列表 ----
        加拿大蒙特利尔 | 货物名称 | ...
    """
    ext = os.path.splitext(excel_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        records = _extract_xlsx(excel_path)
    elif ext == ".xls":
        records = _extract_xls(excel_path)
    else:
        raise ValueError(f"不支持的格式: {ext}")

    # 按 sheet 分组
    sheet_groups: Dict[str, List[Tuple[str, str]]] = {}
    for sheet_name, addr, val in records:
        sheet_groups.setdefault(sheet_name, []).append((addr, val))

    lines = []
    all_values = []

    for sheet_name, cells in sheet_groups.items():
        lines.append(f"\n==== Sheet: {sheet_name} ====")
        for addr, val in cells:
            # 多行值：每行都缩进显示
            val_display = val.replace("\n", "↵ ")
            lines.append(f"[{addr}] {val_display}")
            all_values.append(val.replace("\n", " "))

    # 词汇汇总行：方便大模型做模糊匹配
    lines.append("\n---- 全部词汇（用于模糊匹配）----")
    lines.append(" | ".join(all_values))

    result = "\n".join(lines)

    if output_path:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(result)
        print(f"已写入: {output_path}")

    return result


# ── 命令行入口 ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    EXCEL_PATH = "/home/wzh/project/YiBao/wzh/excel/LS250005-NPO舱单SI&VGM.xls"
    OUTPUT_PATH = "/home/wzh/project/YiBao/wzh/excel/result_text/LS250005_text.txt"

    text = extract_excel_text(EXCEL_PATH, OUTPUT_PATH)
    print(text[:2000])   # 预览前 2000 字符

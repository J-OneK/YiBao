"""
Linux-compatible Excel → Image Converter
依赖: pip install openpyxl xlrd==1.2.0 Pillow fonttools

核心逻辑:
 1. 用 openpyxl(xlsx) / xlrd(xls) 读取表格内容、样式、列宽、行高
 2. 用 Pillow 测量文本实际像素宽度，自动计算 AutoFit 宽/高
 3. 用 Pillow ImageDraw 渲染整张表格为 PNG 图片
 4. 按最大高度分页, 最终裁剪白边
"""

import os
import re
import math
import textwrap
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from PIL import Image, ImageChops, ImageDraw, ImageFont

# ── 常量 ────────────────────────────────────────────────────────────────────
# Excel 默认单位换算: 1 字符宽 ≈ 7 px (at 96dpi)
# Excel ColumnWidth 单位 = "字符数" (正文字体的平均字符宽)
CHAR_WIDTH_PX = 7          # 每字符像素宽
MIN_COL_PX   = 30          # 最小列宽 px
MAX_COL_PX   = 400         # 最大列宽 px (AutoFit 上限)
MIN_ROW_PX   = 16          # 最小行高 px
MAX_ROW_PX   = 300         # 最大行高 px (AutoFit 上限)
H_PAD        = 6           # 单元格水平内边距(每侧)
V_PAD        = 4           # 单元格垂直内边距(每侧)
FONT_SIZE    = 11          # 默认字号 pt
DPI          = 96


# ── 字体加载 ─────────────────────────────────────────────────────────────────
# 脚本所在目录下的 fonts/ 子目录（优先找本地提取的字体）
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_LOCAL_FONTS_DIR = os.path.join(_SCRIPT_DIR, "fonts")


def _load_font(size_pt: int, bold: bool = False, italic: bool = False) -> ImageFont.FreeTypeFont:
    """尝试加载系统中文字体，失败则退回内置字体"""
    candidates = [
        # 优先：脚本旁边 fonts/ 目录下的字体
        os.path.join(_LOCAL_FONTS_DIR, "wqy-zenhei.ttc"),
        os.path.join(_LOCAL_FONTS_DIR, "NotoSansCJK-Regular.otf"),
        os.path.join(_LOCAL_FONTS_DIR, "NotoSansCJK-Regular.ttc"),
        # 系统路径
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    size_px = int(size_pt * DPI / 72)
    for path in candidates:
        if os.path.isfile(path):
            try:
                return ImageFont.truetype(path, size_px)
            except Exception:
                continue
    # 最后 fallback: PIL 内置字体(不支持 CJK, 仅 ASCII)
    return ImageFont.load_default()


_FONT_CACHE: Dict[Tuple, ImageFont.FreeTypeFont] = {}


def get_font(size_pt: int = FONT_SIZE, bold: bool = False) -> ImageFont.FreeTypeFont:
    key = (size_pt, bold)
    if key not in _FONT_CACHE:
        _FONT_CACHE[key] = _load_font(size_pt, bold=bold)
    return _FONT_CACHE[key]


# ── 颜色工具 ─────────────────────────────────────────────────────────────────
def _argb_to_rgb(argb: Optional[str]) -> Optional[Tuple[int, int, int]]:
    """openpyxl 颜色转 (R,G,B). 支持 AARRGGBB 或 RRGGBB 格式."""
    if not argb or argb in ("00000000", "FFFFFFFF", "FF000000"):
        return None
    s = argb.lstrip("#")
    if len(s) == 8:   # AARRGGBB
        r, g, b = int(s[2:4], 16), int(s[4:6], 16), int(s[6:8], 16)
        return (r, g, b)
    if len(s) == 6:   # RRGGBB
        r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
        return (r, g, b)
    return None


def _xl_color(color_obj) -> Optional[Tuple[int, int, int]]:
    """将 openpyxl Color 对象转换成 (R,G,B)"""
    if color_obj is None:
        return None
    try:
        if hasattr(color_obj, "rgb"):
            return _argb_to_rgb(color_obj.rgb)
    except Exception:
        pass
    return None


# ── 数据结构 ──────────────────────────────────────────────────────────────────
@dataclass
class CellStyle:
    bg_color: Optional[Tuple[int, int, int]] = None
    fg_color: Optional[Tuple[int, int, int]] = (0, 0, 0)
    font_size: int = FONT_SIZE
    bold: bool = False
    italic: bool = False
    align_h: str = "left"   # left / center / right
    align_v: str = "top"    # top / middle / bottom
    wrap_text: bool = True
    border_top: bool = False
    border_bottom: bool = False
    border_left: bool = False
    border_right: bool = False


@dataclass
class CellData:
    value: str = ""
    style: CellStyle = field(default_factory=CellStyle)
    merge_span_row: int = 1   # 所跨行数
    merge_span_col: int = 1   # 所跨列数
    is_merged_child: bool = False  # 是合并区域中非左上角的格子


@dataclass
class SheetData:
    name: str = ""
    cells: Dict[Tuple[int, int], CellData] = field(default_factory=dict)  # (row,col) -> CellData  0-indexed
    col_widths_px: List[int] = field(default_factory=list)   # per column px
    row_heights_px: List[int] = field(default_factory=list)  # per row px
    n_rows: int = 0
    n_cols: int = 0


# ── xlsx 解析 (openpyxl) ─────────────────────────────────────────────────────
def _parse_xlsx(path: str) -> List[SheetData]:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue

        # 确定数据范围
        min_r, min_c = ws.min_row or 1, ws.min_column or 1
        max_r, max_c = ws.max_row or 1, ws.max_column or 1
        if ws.max_row is None:
            continue
        n_rows = max_r - min_r + 1
        n_cols = max_c - min_c + 1

        # 合并单元格索引
        merge_map: Dict[Tuple[int,int], Tuple[int,int,int,int]] = {}
        for m in ws.merged_cells.ranges:
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    merge_map[(r, c)] = (m.min_row, m.min_col, m.max_row, m.max_col)

        # 列宽: 先用 Excel 默认值，后续 AutoFit 覆盖
        col_w_px = []
        for ci in range(min_c, max_c + 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            cd = ws.column_dimensions.get(col_letter)
            if cd and cd.width and cd.width > 0:
                col_w_px.append(int(cd.width * CHAR_WIDTH_PX))
            else:
                col_w_px.append(int(8 * CHAR_WIDTH_PX))

        # 行高
        row_h_px = []
        for ri in range(min_r, max_r + 1):
            rd = ws.row_dimensions.get(ri)
            if rd and rd.height and rd.height > 0:
                row_h_px.append(int(rd.height * DPI / 72))
            else:
                row_h_px.append(int(13 * DPI / 72))  # Excel 默认行高 13pt

        # 读取单元格
        cells: Dict[Tuple[int,int], CellData] = {}
        for ri, row in enumerate(ws.iter_rows(min_row=min_r, max_row=max_r,
                                               min_col=min_c, max_col=max_c)):
            for ci, cell in enumerate(row):
                abs_r = ri + min_r
                abs_c = ci + min_c

                # 判断合并
                mr = merge_map.get((abs_r, abs_c))
                is_child = False
                span_r = span_c = 1
                if mr:
                    mr0, mc0, mr1, mc1 = mr
                    is_child = (abs_r != mr0 or abs_c != mc0)
                    span_r = mr1 - mr0 + 1
                    span_c = mc1 - mc0 + 1

                # 样式
                sty = CellStyle()
                try:
                    f = cell.font
                    if f:
                        sty.bold = bool(f.bold)
                        sty.italic = bool(f.italic)
                        if f.size:
                            sty.font_size = int(f.size)
                        if f.color:
                            c = _xl_color(f.color)
                            if c:
                                sty.fg_color = c
                except Exception:
                    pass

                try:
                    fill = cell.fill
                    if fill and fill.fill_type not in (None, "none"):
                        bg = _xl_color(fill.fgColor)
                        if bg:
                            sty.bg_color = bg
                except Exception:
                    pass

                try:
                    aln = cell.alignment
                    if aln:
                        h = (aln.horizontal or "").lower()
                        if h in ("center", "right"):
                            sty.align_h = h
                        elif h in ("general",):
                            sty.align_h = "right" if isinstance(cell.value, (int, float)) else "left"
                        v = (aln.vertical or "").lower()
                        if v in ("center", "bottom"):
                            sty.align_v = "middle" if v == "center" else "bottom"
                        sty.wrap_text = bool(aln.wrapText)
                except Exception:
                    pass

                try:
                    b = cell.border
                    if b:
                        sty.border_top = b.top and b.top.style not in (None, "none", "")
                        sty.border_bottom = b.bottom and b.bottom.style not in (None, "none", "")
                        sty.border_left = b.left and b.left.style not in (None, "none", "")
                        sty.border_right = b.right and b.right.style not in (None, "none", "")
                except Exception:
                    pass

                val = ""
                if cell.value is not None:
                    val = str(cell.value)

                cells[(ri, ci)] = CellData(
                    value=val,
                    style=sty,
                    merge_span_row=span_r,
                    merge_span_col=span_c,
                    is_merged_child=is_child,
                )

        sheet = SheetData(
            name=ws.title,
            cells=cells,
            col_widths_px=col_w_px,
            row_heights_px=row_h_px,
            n_rows=n_rows,
            n_cols=n_cols,
        )
        sheets.append(sheet)

    wb.close()
    return sheets


# ── xls 解析 (xlrd) ──────────────────────────────────────────────────────────
def _parse_xls(path: str) -> List[SheetData]:
    import xlrd

    wb = xlrd.open_workbook(path, formatting_info=True)
    sheets = []

    for si in range(wb.nsheets):
        ws = wb.sheet_by_index(si)
        if ws.visibility != 0:
            continue
        n_rows = ws.nrows
        n_cols = ws.ncols
        if n_rows == 0 or n_cols == 0:
            continue

        # 合并单元格
        merge_map: Dict[Tuple[int,int], Tuple[int,int,int,int]] = {}
        for (r0, r1, c0, c1) in ws.merged_cells:
            for r in range(r0, r1):
                for c in range(c0, c1):
                    merge_map[(r, c)] = (r0, c0, r1-1, c1-1)

        # 列宽 (xlrd 单位 = 1/256 字符)
        col_w_px = []
        for ci in range(n_cols):
            try:
                w256 = ws.colinfo_map[ci].width if ci in ws.colinfo_map else 2048
                col_w_px.append(max(MIN_COL_PX, int(w256 / 256 * CHAR_WIDTH_PX)))
            except Exception:
                col_w_px.append(int(8 * CHAR_WIDTH_PX))

        # 行高 (xlrd 单位 = twips = 1/20 pt)
        row_h_px = []
        for ri in range(n_rows):
            try:
                ht_twips = ws.rowinfo_map[ri].height if ri in ws.rowinfo_map else 255
                ht_pt = ht_twips / 20.0
                row_h_px.append(max(MIN_ROW_PX, int(ht_pt * DPI / 72)))
            except Exception:
                row_h_px.append(int(13 * DPI / 72))

        # 样式表
        xf_list = wb.xf_list
        font_list = wb.font_list
        fmt_map = wb.format_map

        cells: Dict[Tuple[int,int], CellData] = {}
        for ri in range(n_rows):
            for ci in range(n_cols):
                cell = ws.cell(ri, ci)

                mr = merge_map.get((ri, ci))
                is_child = False
                span_r = span_c = 1
                if mr:
                    r0, c0, r1, c1 = mr
                    is_child = (ri != r0 or ci != c0)
                    span_r = r1 - r0 + 1
                    span_c = c1 - c0 + 1

                sty = CellStyle()
                try:
                    xf = xf_list[cell.xf_index]
                    font = font_list[xf.font_index]
                    sty.bold = bool(font.bold)
                    sty.italic = bool(font.italic)
                    if font.height:
                        sty.font_size = int(font.height / 20)   # twips -> pt

                    # 字体颜色
                    try:
                        fc = wb.colour_map.get(font.colour_index)
                        if fc:
                            sty.fg_color = fc[:3]
                    except Exception:
                        pass

                    # 背景色
                    try:
                        bg_idx = xf.background.pattern_colour_index
                        bg = wb.colour_map.get(bg_idx)
                        if bg and bg != (255,255,255) and bg != (0,0,0):
                            sty.bg_color = bg[:3]
                    except Exception:
                        pass

                    # 对齐
                    try:
                        h = xf.alignment.hor_align
                        # xlrd: 0=general,1=left,2=center,3=right,4=fill,5=justify
                        if h == 2:
                            sty.align_h = "center"
                        elif h == 3:
                            sty.align_h = "right"
                        v = xf.alignment.vert_align
                        # 0=top,1=center,2=bottom
                        if v == 1:
                            sty.align_v = "middle"
                        elif v == 2:
                            sty.align_v = "bottom"
                        sty.wrap_text = bool(xf.alignment.text_wrapped)
                    except Exception:
                        pass

                    # 边框
                    try:
                        borders = xf.border
                        sty.border_top = borders.top_line_type > 0
                        sty.border_bottom = borders.bottom_line_type > 0
                        sty.border_left = borders.left_line_type > 0
                        sty.border_right = borders.right_line_type > 0
                    except Exception:
                        pass

                except Exception:
                    pass

                # 值
                val = ""
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    val = cell.value
                elif cell.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_BOOLEAN):
                    num = cell.value
                    val = str(int(num)) if num == int(num) else str(num)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        val = dt.strftime("%Y-%m-%d")
                    except Exception:
                        val = str(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    val = ""
                else:
                    val = str(cell.value) if cell.value else ""

                cells[(ri, ci)] = CellData(
                    value=val,
                    style=sty,
                    merge_span_row=span_r,
                    merge_span_col=span_c,
                    is_merged_child=is_child,
                )

        sheet = SheetData(
            name=ws.name,
            cells=cells,
            col_widths_px=col_w_px,
            row_heights_px=row_h_px,
            n_rows=n_rows,
            n_cols=n_cols,
        )
        sheets.append(sheet)

    return sheets


# ── 裁剪空白区域 ─────────────────────────────────────────────────────────────
def trim_used_range(sheet: SheetData) -> None:
    """
    去掉右侧和下方全空白的行和列，让生成的图片紧凑。
    直接修改 sheet 的 n_rows / n_cols / col_widths_px / row_heights_px。
    """
    # 找最后一个非空格子的行/列
    max_row = -1
    max_col = -1
    for (ri, ci), cd in sheet.cells.items():
        if cd.value and cd.value.strip():
            if ri > max_row:
                max_row = ri
            if ci > max_col:
                max_col = ci

    if max_row < 0:
        # 整张表为空
        sheet.n_rows = 0
        sheet.n_cols = 0
        return

    new_n_rows = max_row + 1
    new_n_cols = max_col + 1

    # 删除超出范围的格子
    to_del = [(r, c) for (r, c) in sheet.cells if r >= new_n_rows or c >= new_n_cols]
    for k in to_del:
        del sheet.cells[k]

    sheet.n_rows = new_n_rows
    sheet.n_cols = new_n_cols
    sheet.col_widths_px = sheet.col_widths_px[:new_n_cols]
    sheet.row_heights_px = sheet.row_heights_px[:new_n_rows]


# ── AutoFit ─────────────────────────────────────────────────────────────────
def measure_text(text: str, font: ImageFont.FreeTypeFont) -> Tuple[int, int]:
    """返回文本 (width_px, height_px)"""
    if not text:
        return 0, int(font.size * 1.2)
    bbox = font.getbbox(text)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def autofit_sheet(sheet: SheetData) -> None:
    """
    对 sheet 的列宽/行高执行 AutoFit:
    - 先算出每个单元格文本所需宽度 → 更新列宽
    - 再根据列宽推算每行所需高度 → 更新行高
    跳过合并格子的子格 (is_merged_child=True).
    """
    # Pass 1: 计算列宽
    best_col_w = list(sheet.col_widths_px)   # 初始值

    for (ri, ci), cd in sheet.cells.items():
        if cd.is_merged_child:
            continue
        if cd.merge_span_col > 1:
            # 跨列合并: 先放到 AutoFit 后再说，暂跳
            continue
        if not cd.value:
            continue
        font = get_font(cd.style.font_size, cd.style.bold)
        # 单行最长
        lines = cd.value.split("\n")
        max_line_w = max((measure_text(ln, font)[0] for ln in lines), default=0)
        needed_w = max_line_w + H_PAD * 2
        needed_w = max(MIN_COL_PX, min(MAX_COL_PX, needed_w))
        if ci < len(best_col_w):
            best_col_w[ci] = max(best_col_w[ci], needed_w)

    sheet.col_widths_px = best_col_w

    # Pass 2: 计算行高
    best_row_h = list(sheet.row_heights_px)

    for (ri, ci), cd in sheet.cells.items():
        if cd.is_merged_child:
            continue
        if cd.merge_span_row > 1:
            continue
        if not cd.value:
            continue
        font = get_font(cd.style.font_size, cd.style.bold)
        cell_w = sheet.col_widths_px[ci] if ci < len(sheet.col_widths_px) else 56

        # 计算实际换行后行数
        lines = cd.value.split("\n")
        total_lines = 0
        _, line_h = measure_text("Ag", font)
        line_h = max(line_h, cd.style.font_size)
        for line in lines:
            if not line:
                total_lines += 1
                continue
            lw, _ = measure_text(line, font)
            usable_w = max(cell_w - H_PAD * 2, 1)
            if lw <= usable_w or not cd.style.wrap_text:
                total_lines += 1
            else:
                total_lines += math.ceil(lw / usable_w)

        total_lines = max(1, total_lines)
        needed_h = total_lines * line_h + V_PAD * 2
        needed_h = max(MIN_ROW_PX, min(MAX_ROW_PX, needed_h))
        if ri < len(best_row_h):
            best_row_h[ri] = max(best_row_h[ri], needed_h)

    sheet.row_heights_px = best_row_h


# ── 渲染 ─────────────────────────────────────────────────────────────────────
def render_sheet_rows(
    sheet: SheetData,
    row_start: int,
    row_end: int,
) -> Image.Image:
    """将 sheet 的 [row_start, row_end] 行渲染为 Pillow Image (0-indexed)"""

    col_xs = [0]
    for w in sheet.col_widths_px:
        col_xs.append(col_xs[-1] + w)
    total_w = col_xs[-1] if col_xs else 1

    row_ys = [0]
    for row_i in range(row_start, row_end + 1):
        h = sheet.row_heights_px[row_i] if row_i < len(sheet.row_heights_px) else MIN_ROW_PX
        row_ys.append(row_ys[-1] + h)
    total_h = row_ys[-1] if row_ys else 1

    img = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # 绘制每个格子
    for ri_abs in range(row_start, row_end + 1):
        ri_local = ri_abs - row_start
        for ci in range(sheet.n_cols):
            cd = sheet.cells.get((ri_abs, ci))
            if cd is None:
                # 空格子画边框
                x0 = col_xs[ci]
                y0 = row_ys[ri_local]
                x1 = col_xs[ci + 1] - 1 if ci + 1 < len(col_xs) else total_w - 1
                y1 = row_ys[ri_local + 1] - 1
                draw.rectangle([x0, y0, x1, y1], outline=(200, 200, 200))
                continue

            if cd.is_merged_child:
                continue

            sty = cd.style

            # 合并区域尺寸
            x0 = col_xs[ci]
            y0 = row_ys[ri_local]
            ci_end = min(ci + cd.merge_span_col, len(col_xs) - 1)
            ri_end_local = min(ri_local + cd.merge_span_row, len(row_ys) - 1)
            x1 = col_xs[ci_end] - 1
            y1 = row_ys[ri_end_local] - 1

            # 背景
            if sty.bg_color:
                draw.rectangle([x0, y0, x1, y1], fill=sty.bg_color)

            # 边框
            border_color = (180, 180, 180)
            if sty.border_top:
                draw.line([(x0, y0), (x1, y0)], fill=(100, 100, 100))
            if sty.border_bottom:
                draw.line([(x0, y1), (x1, y1)], fill=(100, 100, 100))
            if sty.border_left:
                draw.line([(x0, y0), (x0, y1)], fill=(100, 100, 100))
            if sty.border_right:
                draw.line([(x1, y0), (x1, y1)], fill=(100, 100, 100))
            if not (sty.border_top or sty.border_bottom or sty.border_left or sty.border_right):
                draw.rectangle([x0, y0, x1, y1], outline=border_color)

            # 文字
            if cd.value:
                font = get_font(sty.font_size, sty.bold)
                cell_w = x1 - x0
                cell_h = y1 - y0
                fg = sty.fg_color or (0, 0, 0)
                _, line_h = measure_text("Ag", font)
                line_h = max(line_h, sty.font_size) + 2

                # 换行处理
                raw_lines = cd.value.split("\n")
                wrapped_lines = []
                for line in raw_lines:
                    if not line:
                        wrapped_lines.append("")
                        continue
                    lw, _ = measure_text(line, font)
                    usable_w = max(cell_w - H_PAD * 2, 1)
                    if lw <= usable_w or not sty.wrap_text:
                        wrapped_lines.append(line)
                    else:
                        # 按字符逐步拆行
                        buf = ""
                        for ch in line:
                            test = buf + ch
                            tw, _ = measure_text(test, font)
                            if tw > usable_w and buf:
                                wrapped_lines.append(buf)
                                buf = ch
                            else:
                                buf = test
                        if buf:
                            wrapped_lines.append(buf)

                total_text_h = len(wrapped_lines) * line_h

                # 垂直起始坐标
                if sty.align_v == "middle":
                    ty = y0 + max(V_PAD, (cell_h - total_text_h) // 2)
                elif sty.align_v == "bottom":
                    ty = y0 + max(V_PAD, cell_h - total_text_h - V_PAD)
                else:
                    ty = y0 + V_PAD

                for line in wrapped_lines:
                    if not line:
                        ty += line_h
                        continue
                    lw, _ = measure_text(line, font)
                    if sty.align_h == "center":
                        tx = x0 + max(H_PAD, (cell_w - lw) // 2)
                    elif sty.align_h == "right":
                        tx = x0 + max(H_PAD, cell_w - lw - H_PAD)
                    else:
                        tx = x0 + H_PAD
                    if ty + line_h > y1:
                        break
                    draw.text((tx, ty), line, font=font, fill=fg)
                    ty += line_h

    return img


def trim_white_border(image: Image.Image, diff_threshold: int = 10, padding: int = 2) -> Image.Image:
    if image.mode != "RGB":
        image = image.convert("RGB")
    bg = Image.new("RGB", image.size, (255, 255, 255))
    diff = ImageChops.difference(image, bg).convert("L")
    diff = diff.point(lambda x: 255 if x > diff_threshold else 0)
    bbox = diff.getbbox()
    if not bbox:
        return image
    l, t, r, b = bbox
    return image.crop((
        max(l - padding, 0),
        max(t - padding, 0),
        min(r + padding, image.width),
        min(b + padding, image.height),
    ))


# ── 主入口 ────────────────────────────────────────────────────────────────────
def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name)


def excel_to_images_linux(
    excel_path: str,
    output_dir: str,
    max_height_px: int = 1600,
    max_width_px: int = 2400,
    trim_white: bool = True,
    diff_threshold: int = 10,
    padding: int = 2,
) -> None:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(excel_path)
    os.makedirs(output_dir, exist_ok=True)

    ext = os.path.splitext(excel_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        sheets = _parse_xlsx(excel_path)
    elif ext in (".xls",):
        sheets = _parse_xls(excel_path)
    else:
        raise ValueError(f"Unsupported format: {ext}")

    base_name = sanitize_filename(os.path.splitext(os.path.basename(excel_path))[0])
    image_index = 1

    for sheet in sheets:
        if sheet.n_rows == 0 or sheet.n_cols == 0:
            continue

        # 先裁剪空白行列
        trim_used_range(sheet)
        if sheet.n_rows == 0 or sheet.n_cols == 0:
            print(f"  -> Skip empty sheet: {sheet.name}")
            continue

        print(f"Processing sheet: {sheet.name}  ({sheet.n_rows} rows × {sheet.n_cols} cols)")

        # AutoFit
        autofit_sheet(sheet)

        # 按最大高度分页
        chunks = []
        cur_start = 0
        cur_h = 0
        for ri in range(sheet.n_rows):
            rh = sheet.row_heights_px[ri] if ri < len(sheet.row_heights_px) else MIN_ROW_PX
            if cur_h + rh > max_height_px and ri > cur_start:
                chunks.append((cur_start, ri - 1))
                cur_start = ri
                cur_h = 0
            cur_h += rh
        if cur_start < sheet.n_rows:
            chunks.append((cur_start, sheet.n_rows - 1))

        for row_start, row_end in chunks:
            img = render_sheet_rows(sheet, row_start, row_end)

            # 限制最大宽度
            if img.width > max_width_px:
                ratio = max_width_px / img.width
                img = img.resize(
                    (max_width_px, int(img.height * ratio)),
                    Image.LANCZOS,
                )

            if trim_white:
                img = trim_white_border(img, diff_threshold=diff_threshold, padding=padding)

            filename = f"{base_name}_{image_index}.png"
            output_path = os.path.join(output_dir, filename)
            img.save(output_path)
            print(f"  -> Saved {output_path}  ({img.width}×{img.height})")
            image_index += 1


# ── 配置 & 运行 ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    EXCEL_PATH = "/home/wzh/project/YiBao/wzh/excel/LS250005-NPO舱单SI&VGM.xls"
    OUTPUT_DIR = "/home/wzh/project/YiBao/wzh/excel/result_imgs_linux"

    excel_to_images_linux(
        excel_path=EXCEL_PATH,
        output_dir=OUTPUT_DIR,
        max_height_px=1600,
        max_width_px=2400,
        trim_white=False,
        diff_threshold=10,
        padding=2,
    )
